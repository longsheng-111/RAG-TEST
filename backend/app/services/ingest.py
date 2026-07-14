"""
文件处理与文本切片模块
"""
import base64
import re
import shutil
from pathlib import Path
from typing import List

from PyPDF2 import PdfReader
import fitz  # PyMuPDF
from docx import Document
from openpyxl import load_workbook

from app.core.config import settings


# ============================================================
#  文件读取
# ============================================================

def read_text_from_file(file_path: Path) -> str:
    """
    根据文件后缀选择合适的读取方式提取纯文本。
    支持: .txt / .md / .csv / .json / .log / .pdf / .docx / .xlsx / .xlsm
    """
    suffix = file_path.suffix.lower()

    # ---- 纯文本类 ----
    if suffix in {".txt", ".md", ".csv", ".json", ".log"}:
        for enc in ("utf-8", "utf-16", "gbk"):
            try:
                return file_path.read_text(encoding=enc)
            except (UnicodeDecodeError, UnicodeError):
                continue
        return file_path.read_text(encoding="gbk", errors="ignore")

    # ---- PDF ----
    if suffix == ".pdf":
        reader = PdfReader(str(file_path))
        text = "\n".join(
            (page.extract_text() or "") for page in reader.pages
        )
        # 如果提取不到文字（图片型PDF），走视觉模型
        if not text.strip():
            text = extract_text_with_qwen_vl(str(file_path))
        return text

    # ---- Word ----
    if suffix == ".docx":
        doc = Document(str(file_path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    # ---- Excel ----
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        wb = load_workbook(str(file_path), data_only=True)
        parts: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join(
                    str(cell) for cell in row if cell is not None
                )
                if row_text.strip():
                    parts.append(row_text)
        return "\n".join(parts)

    raise ValueError(f"不支持的文件格式: {suffix}")


# ============================================================
#  图片型 PDF 处理（通义千问视觉模型）
# ============================================================

def extract_text_with_qwen_vl(pdf_path: str) -> str:
    """
    使用 Qwen3.7-Plus 视觉模型识别图片型 PDF 中的文字。
    逐页渲染为 JPEG → base64 → 调用 OpenAI 兼容 API。
    """
    from openai import OpenAI

    api_key = settings.dashscope_api_key
    base_url = settings.dashscope_base_url

    if not api_key:
        raise RuntimeError("未配置 DASHSCOPE_API_KEY，无法处理图片型 PDF")

    client = OpenAI(api_key=api_key, base_url=base_url)
    doc = fitz.open(pdf_path)
    full_text: list[str] = []

    try:
        for page_num in range(len(doc)):
            page = doc[page_num]

            # 页面渲染为 JPEG
            pix = page.get_pixmap(dpi=200)
            img_bytes = pix.tobytes("jpg")
            img_b64 = base64.b64encode(img_bytes).decode("utf-8")

            # 调用 Qwen3.7-Plus 视觉模型 (OpenAI 兼容)
            try:
                response = client.chat.completions.create(
                    model="qwen-plus",
                    messages=[{
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"},
                            },
                            {
                                "type": "text",
                                "text": "请提取图片中的所有文字，保持原有格式和段落结构。",
                            },
                        ],
                    }],
                    max_tokens=4096,
                )
                content = response.choices[0].message.content
                if content:
                    full_text.append(content)
            except Exception as e:
                print(f"[Qwen-VL] 第 {page_num + 1} 页失败: {e}")

    finally:
        doc.close()

    return "\n\n".join(full_text).strip()


# ============================================================
#  文本切片
# ============================================================

def split_text(text: str, source_file: str = "") -> list[str]:
    """
    文本切片主入口：清洗 → 标题切分 → 递归字符切分
    """
    # 清洗空行
    cleaned_lines = [line.strip() for line in text.splitlines() if line.strip()]
    cleaned = "\n".join(cleaned_lines)
    if not cleaned:
        return []

    return _split_by_paragraphs(cleaned, source_file)


def _split_by_paragraphs(text: str, source_file: str = "") -> list[str]:
    """按段落 / 标题切分文本"""
    header_chunks = _split_by_headers(text)

    final_chunks: list[str] = []
    for chunk in header_chunks:
        content = chunk.strip()
        if not content:
            continue
        if len(content) <= settings.max_chunk_size:
            final_chunks.append(content)
            continue

        # 超长标题块：把标题路径前缀保留在每个子块中，避免答案被切碎后丢失上下文
        lines = content.split("\n", 1)
        if len(lines) == 2 and lines[0].strip():
            prefix, body = lines[0].strip(), lines[1].strip()
            sub_chunks = _recursive_split(body)
            for sub in sub_chunks:
                final_chunks.append(f"{prefix}\n\n{sub}")
        else:
            final_chunks.extend(_recursive_split(content))

    return final_chunks


def _split_by_headers(text: str) -> list[str]:
    """Markdown 标题切分：保留标题层级作为路径前缀"""
    from langchain_text_splitters import MarkdownHeaderTextSplitter

    headers_to_split = [
        ("#", "大章节"),
        ("##", "小节"),
        ("###", "小点"),
        ("####", "段落"),
    ]

    # 去重标题（按出现顺序）
    seen: set[str] = set()
    unique_headers: list[tuple[str, str]] = []
    for h in headers_to_split:
        if h[1] not in seen:
            seen.add(h[1])
            unique_headers.append(h)

    splitter = MarkdownHeaderTextSplitter(
        headers_to_split_on=unique_headers,
        strip_headers=True,
    )
    docs = splitter.split_text(text)

    chunks: list[str] = []
    for doc in docs:
        content = doc.page_content.strip()
        if not content:
            continue

        # 构建标题路径前缀
        header_parts: list[str] = []
        for level in [h[1] for h in unique_headers]:
            if level in doc.metadata and doc.metadata[level]:
                header_parts.append(doc.metadata[level])

        if header_parts:
            prefix = " > ".join(header_parts)
            content = f"{prefix}\n\n{content}"

        chunks.append(content)

    return chunks


def _recursive_split(text: str) -> list[str]:
    """递归字符切分：按分隔符优先级逐步切分"""
    separators = ["\n\n", "\n", "。", "！", "？", "；", "，", " ", ""]
    return _split_recursive(text, separators, settings.max_chunk_size, settings.chunk_overlap)


def _split_recursive(text: str, separators: list[str], chunk_size: int, chunk_overlap: int) -> list[str]:
    """递归切分核心"""
    # 如果文本已经够短，直接返回
    if len(text) <= chunk_size:
        return [text] if text.strip() else []

    # 尝试用当前分隔符切
    sep = separators[0] if separators else ""
    if sep:
        segments = text.split(sep)
    else:
        # 最后手段：按字符切
        segments = list(text)

    chunks: list[str] = []
    current = ""

    for seg in segments:
        candidate = current + (sep if current and sep else "") + seg

        if len(candidate) <= chunk_size:
            current = candidate
        else:
            # 当前 chunk 已满
            if current.strip():
                chunks.append(current)

            # 处理当前段落
            if len(seg) > chunk_size:
                # 段落太长，降级使用下一个分隔符
                if len(separators) > 1:
                    chunks.extend(
                        _split_recursive(seg, separators[1:], chunk_size, chunk_overlap)
                    )
                else:
                    # 无分隔符可用，硬切
                    for i in range(0, len(seg), chunk_size - chunk_overlap):
                        sub = seg[i:i + chunk_size]
                        if sub.strip():
                            chunks.append(sub)
                current = ""
            else:
                current = seg

    if current.strip():
        chunks.append(current)

    return chunks


# ============================================================
#  完整处理流水线
# ============================================================

def process_file(file_path: str, collection_name: str) -> dict:
    """
    完整文件处理流水线：读取 → 切片 → 嵌入 → 入库
    返回处理的 chunk 信息
    """
    from app.core.vector_store import add_texts

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    # 1. 读取文本
    raw_text = read_text_from_file(path)

    if not raw_text.strip():
        raise ValueError(f"文件内容为空: {path.name}")

    # 2. 切分
    chunks = split_text(raw_text, path.name)

    if not chunks:
        raise ValueError(f"切分结果为空: {path.name}")

    # 3. 构建元数据并入库
    metadatas = [
        {
            "file_name": path.name,
            "source": str(path),
            "chunk_index": i,
            "collection_name": collection_name,
        }
        for i in range(len(chunks))
    ]

    count = add_texts(chunks, metadatas, collection_name)

    return {
        "file_name": path.name,
        "chunks": count,
        "collection_name": collection_name,
    }
